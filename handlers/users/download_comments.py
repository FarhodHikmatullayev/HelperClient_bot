import datetime
import tempfile

from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import CallbackQuery, Message
import openpyxl
import os

from data.config import ADMINS
from keyboards.default.menu import back_menu_keyboard
from loader import dp, db, bot


async def download_all_comments_function():
    comments = await db.select_all_comments()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = 'BAHOLAGAN SHAXS'
    worksheet['C1'] = 'FILIAL NOMI'
    # worksheet['D1'] = 'LAVOZIM'
    worksheet['D1'] = 'XODIM ID RAQAMI'
    worksheet['E1'] = 'XODIM ISM FAMILIYASI'
    worksheet['F1'] = 'BAHO'
    worksheet['G1'] = 'VAQT'
    worksheet['H1'] = 'FIKR'

    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='BAHOLAGAN SHAXS')
    worksheet.cell(row=1, column=3, value='FILIAL NOMI')
    # worksheet.cell(row=1, column=4, value='LAVOZIM')
    worksheet.cell(row=1, column=4, value='XODIM ID RAQAMI')
    worksheet.cell(row=1, column=5, value='XODIM ISM FAMILIYASI')
    worksheet.cell(row=1, column=6, value='BAHO')
    worksheet.cell(row=1, column=7, value='VAQT')
    worksheet.cell(row=1, column=8, value='FIKR')
    tr = 0
    for row, comment in enumerate(comments, start=2):
        filial_id = comment['branch_id']
        department_id = comment['department_id']
        if department_id:
            department = await db.select_department(id=department_id)
            department_name = department['name']
        else:
            department_name = '-'

        user_id = comment['user_id']

        branch = await db.select_branch(id=filial_id)
        users = await db.select_users(id=user_id)
        user = users[0]

        branch_name = branch['name']
        full_name = user['full_name']
        username = user['username']
        phone = user['phone']
        telegram_id = user['telegram_id']
        code = str(comment['employee_code'])
        employees = await db.select_employee(code=code)
        if employees:
            employee = employees[0]
            employee_name = employee['full_name']
        else:
            employee_name = '-'
        print('code', code)
        print('type(code)', type(code))
        if code == "None":
            code = '-'
            print('bajarildi')

        tr += 1
        worksheet.cell(row=row, column=1, value=tr)
        worksheet.cell(row=row, column=2, value=full_name)
        worksheet.cell(row=row, column=3, value=branch_name)
        # worksheet.cell(row=row, column=4, value=department_name)
        worksheet.cell(row=row, column=4, value=code)
        worksheet.cell(row=row, column=5, value=employee_name)
        worksheet.cell(row=row, column=6, value=comment['mark'])
        worksheet.cell(row=row, column=7,
                       value=(comment['created_at'] + datetime.timedelta(hours=5)).strftime('%d.%m.%Y %H:%M'))
        worksheet.cell(row=row, column=8, value=comment['message'])

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Comments_data.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(Command('download_all_comments'), user_id=ADMINS, state='*')
async def download_emp(message: Message, state: FSMContext):
    await state.finish()
    temp_dir = await download_all_comments_function()

    with open(os.path.join(temp_dir, 'Comments_data.xlsx'), 'rb') as file:
        await message.answer_document(document=file)
        # await bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)

    os.remove(os.path.join(temp_dir, 'Comments_data.xlsx'))


@dp.message_handler(Command('download_all_comments'), state='*')
async def download_emp(message: Message, state: FSMContext):
    await state.finish()
    text = "Bu komanda faqat adminlar uchun"
    await message.answer(text=text, reply_markup=back_menu_keyboard)
