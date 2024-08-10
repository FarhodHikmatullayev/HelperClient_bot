import tempfile
from typing import Union

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import CallbackQuery, Message
import openpyxl
import os

from data.config import ADMINS
from keyboards.default.menu import back_menu_keyboard
from loader import dp, db, bot


async def download_all_employees_function():
    employees = await db.select_all_employees()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = 'ISM FAMILIYA'
    worksheet['C1'] = 'LAVOZIM'
    worksheet['D1'] = 'FILIAL'
    worksheet['E1'] = 'ID'

    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='ISM FAMILIYA')
    worksheet.cell(row=1, column=3, value="LAVOZIM")
    worksheet.cell(row=1, column=4, value='FILIAL')
    worksheet.cell(row=1, column=5, value='ID')
    tr = 0
    for row, employee in enumerate(employees, start=2):
        filial_id = employee['filial_id']
        department_id = employee['department_id']

        branch = await db.select_branch(id=filial_id)
        department = await db.select_department(id=department_id)

        branch_name = branch['name']
        department_name = department['name']

        tr += 1
        worksheet.cell(row=row, column=1, value=tr)
        worksheet.cell(row=row, column=2, value=employee['full_name'])
        worksheet.cell(row=row, column=3, value=department_name)
        worksheet.cell(row=row, column=4, value=branch_name)
        worksheet.cell(row=row, column=5, value=employee['code'])

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Employees_data.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(Command('download_employees'), user_id=ADMINS, state='*')
async def download_emp(message: Message, state: FSMContext):
    await state.finish()
    temp_dir = await download_all_employees_function()

    with open(os.path.join(temp_dir, 'Employees_data.xlsx'), 'rb') as file:
        await message.answer_document(document=file)
        # await bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)

    os.remove(os.path.join(temp_dir, 'Employees_data.xlsx'))


@dp.message_handler(Command('download_employees'), state='*')
async def download_emp(message: Message, state: FSMContext):
    await state.finish()
    text = "Bu komanda faqat adminlar uchun"
    await message.answer(text=text, reply_markup=back_menu_keyboard)
