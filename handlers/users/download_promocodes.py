import tempfile
from typing import Union

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import CallbackQuery, Message
import openpyxl
import os

from data.config import ADMINS
from keyboards.default.menu import back_menu_keyboard
from loader import dp, db, bot


async def download_all_promo_codes():
    promo_codes = await db.select_all_promo_codes()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet['A1'] = 'T/r'
    worksheet['B1'] = 'PROMOCODE'
    worksheet['C1'] = 'FULL_NAME'
    worksheet['D1'] = 'USERNAME'
    worksheet['E1'] = 'PHONE'
    worksheet['F1'] = 'TELEGRAM ID'
    worksheet['G1'] = 'TIME'

    worksheet.cell(row=1, column=1, value='№')
    worksheet.cell(row=1, column=2, value='PROMOCODE')
    worksheet.cell(row=1, column=3, value='FULL_NAME')
    worksheet.cell(row=1, column=4, value="USERNAME")
    worksheet.cell(row=1, column=5, value='PHONE')
    worksheet.cell(row=1, column=6, value='TELEGRAM ID')
    worksheet.cell(row=1, column=7, value='TIME')
    tr = 0
    for row, promo_code in enumerate(promo_codes, start=2):
        user_id = promo_code['user_id']

        users = await db.select_users(id=user_id)
        user = users[0]

        full_name = user['full_name']
        username = user['username']
        phone = user['phone']
        telegram_id = user['telegram_id']

        tr += 1
        worksheet.cell(row=row, column=1, value=tr)
        worksheet.cell(row=row, column=2, value=promo_code['promocode'])
        worksheet.cell(row=row, column=3, value=full_name)
        worksheet.cell(row=row, column=4, value=username)
        worksheet.cell(row=row, column=5, value=phone)
        worksheet.cell(row=row, column=6, value=telegram_id)
        worksheet.cell(row=row, column=7, value=str(promo_code['created_at']))

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, 'Promo_codes_data.xlsx')
    workbook.save(file_path)

    return temp_dir


@dp.message_handler(Command('download_all_promo_codes'), user_id=ADMINS, state='*')
async def download_promo(message: Message, state: FSMContext):
    await state.finish()
    temp_dir = await download_all_promo_codes()

    with open(os.path.join(temp_dir, 'Promo_codes_data.xlsx'), 'rb') as file:
        await message.answer_document(document=file)
        # await bot.delete_message(chat_id=message.chat.id, message_id=message.message_id)

    os.remove(os.path.join(temp_dir, 'Promo_codes_data.xlsx'))


@dp.message_handler(Command('download_all_promo_codes'), state='*')
async def download_promo(message: Message, state: FSMContext):
    await state.finish()
    text = "Bu komanda faqat adminlar uchun"
    await message.answer(text=text, reply_markup=back_menu_keyboard)
