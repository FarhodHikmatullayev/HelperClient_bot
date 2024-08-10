import os

import psycopg2
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.types import ContentType, ContentTypes
from openpyxl.reader.excel import load_workbook

from data.config import ADMINS, SUPERADMINS
from keyboards.default.menu import back_menu_keyboard
from loader import dp, db
from states.add_employee import AddEmployeeState


@dp.message_handler(Command('upload_employees'), user_id=ADMINS, state='*')
async def get_employees_excel_file(message: types.Message, state: FSMContext):
    await state.finish()
    text = "Xodimlar jadvalini yuboring (Excel)"
    await message.answer(text=text)
    await AddEmployeeState.waiting_for_excel_file.set()


@dp.message_handler(Command('upload_employees'), state='*')
async def get_employees_excel_file(message: types.Message, state: FSMContext):
    await state.finish()
    text = "Bu komanda faqat adminlar uchun"
    await message.answer(text=text, reply_markup=back_menu_keyboard)


@dp.message_handler(content_types=[ContentType.DOCUMENT], state=AddEmployeeState.waiting_for_excel_file)
async def save_employees_from_excel(message: types.Message, state: FSMContext):
    if message.document.file_name.endswith('.xlsx') or message.document.file_name.endswith('.xls'):
        file_id = message.document.file_id
        file_name = message.document.file_name
        file_path = os.path.join('/tmp', file_name)

        await message.document.download(file_path)
        await db.delete_all_employees()
        employees = await db.select_all_employees()
        try:
            # Load the Excel file
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # Insert data from the Excel file into the "employee" table
            for row in range(2, worksheet.max_row + 1):
                full_name = worksheet.cell(row=row, column=2).value
                department_name = worksheet.cell(row=row, column=3).value
                branch_name = worksheet.cell(row=row, column=4).value
                code = worksheet.cell(row=row, column=5).value

                if not all([full_name, department_name, branch_name, code]):
                    break
                departments = await db.select_departmetns(name=department_name)
                branches = await db.select_branches(name=branch_name)
                if departments:
                    department = departments[0]
                    department_id = department['id']
                else:
                    department = await db.create_department(name=department_name)
                    department_id = department['id']

                if branches:
                    branch = branches[0]
                    branch_id = branch['id']
                else:
                    branch = await db.create_branch(name=branch_name)
                    branch_id = branch['id']
                branch_department = await db.select_department_filial(department_id=department_id, filial_id=branch_id)
                if not branch_department:
                    branch_department = await db.create_department_filial(department_id=department_id,
                                                                          filial_id=branch_id)
                employee = await db.create_employee(full_name=full_name, department_id=department_id,
                                                    filial_id=branch_id, code=str(code))

                await state.finish()

            await message.answer(f'"{file_name}" dagi ma\'lumotlar muvaffaqiyatli saqlandi')
        except (Exception, psycopg2.Error) as error:
            await message.answer(f'Error occurred while processing the file: {error}\n'
                                 f'Kiritgan faylingizdagi ma\'lumotlarda xatolik mavjud,\n'
                                 f'Iltimos, o\'zgartirib qayta jo\'nating.')
            await AddEmployeeState.waiting_for_excel_file.set()
        finally:
            os.remove(file_path)

    else:
        await message.answer("Bu fayl excel fayl emas. Iltimos faqat excel fayl kiriting.")
        await AddEmployeeState.waiting_for_excel_file.set()


@dp.message_handler(state=AddEmployeeState.waiting_for_excel_file, content_types=ContentTypes.ANY)
async def save_employees_from_excel(message: types.Message):
    text = "Siz excel file kiritmadingiz, iltimos excel file kiriting"
    await message.answer(text=text)
    await AddEmployeeState.waiting_for_excel_file.set()
